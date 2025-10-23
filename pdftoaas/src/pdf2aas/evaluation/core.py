"""Abstract evaluation class and common evaluation functions."""
# ruff: noqa: ERA001 # Allow commented out code for now

import logging
import math
import re
from collections import Counter
from collections.abc import Callable
from pathlib import Path
from typing import Any, ClassVar, Literal

import matplotlib.pyplot as plt
import openpyxl
from Levenshtein import distance
from matplotlib.container import BarContainer

from pdf2aas.model import Property, PropertyDefinition, SimplePropertyDataType

from .article import EvaluationArticle
from .counts import EvaluationCounts
from .prompt import EvaluationPrompt
from .values import EvaluationValues

logger = logging.getLogger(__name__)


def _plot_discrete_hist(
    data: Any,
    limit: int = 0,
    **kwargs: Any,
) -> tuple[BarContainer, dict[Any, int]]:
    counts: dict[Any, int] = Counter(data)
    counts = dict(sorted(counts.items(), key=lambda item: item[1], reverse=True))
    if limit > 0:
        counts = dict(list(counts.items())[0:limit])
    bars = plt.bar(list(counts.keys()), list(counts.values()), **kwargs)
    plt.xticks(rotation=90)
    plt.yticks()
    plt.bar_label(bars)
    plt.grid(axis="y")
    return bars, counts


def _convert_to_float(number: Any) -> float:
    if isinstance(number, float):
        return number
    if not isinstance(number, str):
        return float(number)

    if "," in number and "." in number:
        if number.rfind(",") > number.rfind("."):
            decimal_separator = ","
            thousand_separator = "."
        else:
            decimal_separator = "."
            thousand_separator = ","
    elif "," in number:
        decimal_separator = ","
        thousand_separator = None
    elif "." in number:
        decimal_separator = "."
        thousand_separator = None
    else:
        return float(number)
    if thousand_separator:
        number = number.replace(thousand_separator, "")
    number = number.replace(decimal_separator, ".")
    return float(number)


class Evaluation:
    """Base class for evaluations of the pdf2aas conversion.

    Attributes:
        ignore_properties (ClassVar[set[str]]): Set of property names to be
            ignored during evaluation.
        float_tolerance (float): Tolerance level for comparing floating-point numbers,
            default is 0.01.
        char_tolerance (int): Tolerance level for character differences (levensthein distance),
            default is 2.
        case_sensitive (bool): Flag indicating if the evaluation should be case sensitive,
            default is False.
        true_values (ClassVar[list[str]]): List of string representations
            considered as true values.
        false_values (ClassVar[list[str]]): List of string representations
            considered as false values.
        ignore_expected_values (ClassVar[list]): Don't evaluate properties,
            that have a value in this list. Usefull to exclude empty properties from the evaluation,
            including None, empty string, and "MIN" on default.
        value_datasheet_regex (dict): Dictionary mapping prpperty definition ids (e.g. eclass) to
            regular expressions used to search the datasheet instead of using the value.
        equal_str_values (dict): Dictionary for storing mappings of equivalent string values.
        table_header (ClassVar[list[str]]): Column names for the evaluation table.
        definitions_table_header (ClassVar[list[str]]): Column names for the definitions table.

    """

    ignore_properties: ClassVar[set[str]] = set()
    float_tolerance = 1e-2
    char_tolerance = 2
    case_sensitive = False
    true_values: ClassVar[list[str]] = ["y", "yes", "t", "true", "on", "1"]
    false_values: ClassVar[list[str]] = ["n", "no", "f", "false", "off", "0"]
    ignore_expected_values: ClassVar[list] = [None, ""]
    value_datasheet_regex: ClassVar[dict] = {}
    equal_str_values: ClassVar[dict] = {}

    table_header: ClassVar[list[str]] = [
        "id",
        "name",
        "extracted",
        "expected",
        "difference",
        "correct",
        "similar",
        "different",
        "article",
        "unit",
        "submodel",
    ]
    definitions_table_header: ClassVar[list[str]] = [
        "article",
        "id",
        "name",
        "type",
        "unit",
        "definition",
        "values",
    ]

    def __init__(self) -> None:
        """Initialize the Evaluation with empty and default values."""
        self.articles: list[EvaluationArticle] = []
        self.definitions: dict[str, PropertyDefinition] = {}
        self.extracted_properties: dict[str, list[Property]] = {}
        self.counts: dict[str, EvaluationCounts] = {}
        self.counts_sum: EvaluationCounts = EvaluationCounts()
        self.values: dict[str, EvaluationValues] = {}
        self.ignored_values: dict[str, EvaluationValues] = {}
        self.prompts: list[EvaluationPrompt] = []
        self.language = "en"

    def _compare_string(self, value: str, expected: str) -> tuple[float, bool]:
        if value in self.equal_str_values and self.equal_str_values[value] == expected:
            return 0, False
        if not self.case_sensitive:
            value = value.lower()
            expected = expected.lower()
        try:
            difference = distance(value, expected)
        except ValueError:
            pass
        else:
            similar = (
                difference <= self.char_tolerance
                and difference != 0
                and len(expected) > self.char_tolerance
            )
            return difference, similar
        return -1, False

    def _compare_numeric(self, value: Any, expected: Any) -> tuple[float, bool]:
        if value is None:
            value = 0
        if expected is None:
            expected = 0
        try:
            expected = _convert_to_float(expected)
            difference = abs(expected - float(value))
            similar = math.isclose(expected, float(value), rel_tol=self.float_tolerance)
        except (ValueError, TypeError):
            pass
        else:
            return difference, similar and difference != 0
        return -1, False

    def _compare_bool(self, value: Any, expected: Any) -> tuple[float, bool]:
        if not isinstance(value, bool):
            if str(value).lower() in self.true_values:
                value = True
            elif str(value).lower() in self.false_values:
                value = False
            else:
                return -1, False
        if not isinstance(expected, bool):
            if str(expected).lower() in self.true_values:
                expected = True
            elif str(expected).lower() in self.false_values:
                expected = False
            else:
                return -1, False
        return 0 if value == expected else 1, False

    def _compare(
        self,
        value: Any,
        expected: Any,
        type_str: SimplePropertyDataType,
    ) -> tuple[float, bool]:
        if expected == value:
            return 0, False

        if value is None or expected is None:
            if str(value).strip() == str(expected).strip():
                return 0, False
            return -1, False

        result = -1.0, False
        if type_str == "string":
            result = self._compare_string(str(value), str(expected))
        elif type_str == "numeric":
            result = self._compare_numeric(value, expected)
        elif type_str == "bool":
            result = self._compare_bool(value, expected)
        return result

    def _check_ignored_expected(
        self,
        id_: str,
        property_: Property,
        definition: PropertyDefinition,
        article: EvaluationArticle,
    ) -> tuple[bool, Any]:
        ignored = False
        expected = None
        if property_.label in self.ignore_properties:
            ignored = True
            logger.debug(
                "Property %s (%s) ignored for article: %s",
                id_,
                property_.label,
                article.name,
            )

        expected = article.values.get(id_)

        if definition.type == "numeric":
            try:
                expected = _convert_to_float(expected)
            except (ValueError, TypeError):
                ignored = True
                logger.debug(
                    "Property %s (%s) ignored because of expected value (%s) for article: %s",
                    id_,
                    property_.label,
                    expected,
                    article.name,
                )
        if expected in self.ignore_expected_values:
            ignored = True
            logger.debug(
                "Property %s (%s) ignored because of expected value (%s) for article: %s",
                id_,
                property_.label,
                expected,
                article.name,
            )

        if article.datasheet_text and id_ in self.value_datasheet_regex:
            label = self.value_datasheet_regex.get(id_, "")
            expected_in_datasheet = re.search(
                r"^" + label + r"\s([0-9,.]+)\s(mm|g)",
                article.datasheet_text,
                re.IGNORECASE | re.MULTILINE,
            )
            if expected_in_datasheet:
                expected = expected_in_datasheet.group(1)
            else:
                ignored = True
                logger.debug(
                    "Property %s (%s) ignored because of expected value (%s) not found in preprocessed datasheet for article: %s",  # noqa: E501
                    id_,
                    property_.label,
                    expected,
                    article.name,
                )
        return ignored, expected

    def _calc_property_counts(
        self,
        article: EvaluationArticle,
        property_: Property,
    ) -> None:
        id_, definition = self._get_id_definition(property_)

        if id_ not in self.counts:
            self.counts[id_] = EvaluationCounts()
        counts = self.counts[id_]
        counts.extracted += 1

        if (definition is None) or (id_ not in article.values):
            counts.extra += 1
            logger.debug(
                "Found extra property %s (%s) for article: %s",
                id_,
                property_.label,
                article.name,
            )
            return

        ignored, expected = self._check_ignored_expected(
            id_,
            property_,
            definition,
            article,
        )
        value = property_.value
        difference, similar = self._compare(value, expected, definition.type)

        if ignored:
            counts.ignored += 1
            if id_ not in self.ignored_values:
                self.ignored_values[id_] = EvaluationValues()
            values = self.ignored_values[id_]
        else:
            if id_ not in self.values:
                self.values[id_] = EvaluationValues()
            values = self.values[id_]

            if value is not None:
                counts.value += 1
            if difference == 0:
                counts.correct += 1
            elif similar:
                logger.debug(
                    "Property %s (%s) value %s is similar to expected %s for article: %s",
                    id_,
                    property_.label,
                    value,
                    expected,
                    article.name,
                )
                counts.similar += 1
            else:
                counts.different += 1

        values.difference.append(difference)
        values.extracted.append(value)
        values.expected.append(expected)
        values.similar.append(similar)
        values.articles.append(article.name)
        values.unit.append(property_.unit if property_.unit else "")
        values.submodel.append(article.values.get(id_))

    def _get_id_definition(self, property_: Property) -> tuple[str, PropertyDefinition | None]:
        if property_.definition is not None and property_.definition_id is not None:
            id_ = property_.definition_id
            definition = self.definitions.get(id_)
            if definition is None:
                definition = property_.definition
                self.definitions[id_] = property_.definition
            return id_, definition
        return property_.id, None

    def evaluate(self) -> None:
        """Evaluate the defined and extracted values of all articles.

        Resets counts and values etc. for plotting.
        """
        self.counts = {}
        self.values = {}
        self.ignored_values = {}
        self.counts_sum = EvaluationCounts()
        if len(self.extracted_properties) == 0:
            return
        for article in self.articles:
            extracted_properties = self.extracted_properties.get(article.name)
            if extracted_properties is None:
                logger.warning("No extracted properties for article: %s", article.name)
                continue
            for property_ in extracted_properties:
                self._calc_property_counts(article, property_)

        for counts in self.counts.values():
            self.counts_sum.extracted += counts.extracted
            self.counts_sum.ignored += counts.ignored
            self.counts_sum.extra += counts.extra
            self.counts_sum.value += counts.value
            self.counts_sum.correct += counts.correct
            self.counts_sum.similar += counts.similar
            self.counts_sum.different += counts.different

    def _create_definitions_table(self) -> list[list[str]]:
        return [
            [
                article.name,
                definition.id,
                definition.get_name(self.language, ""),
                definition.type,
                definition.unit,
                definition.get_definition(self.language, ""),
                str(definition.values),
            ]
            for article in self.articles
            for definition in article.definitions
        ]

    def _create_table(
        self,
        considered_values: dict[str, EvaluationValues] | None = None,
        articles: list[str] | None = None,
    ) -> list[list[str]]:
        rows = []
        if considered_values is None:
            considered_values = self.values
        for property_id, values in considered_values.items():
            for i in range(len(values.extracted)):
                if articles and values.articles[i] not in articles:
                    continue
                row = [
                    property_id,
                    self.definitions[property_id].get_name(self.language),
                    str(values.extracted[i])
                    if isinstance(values.extracted[i], list)
                    else values.extracted[i],
                    values.expected[i],
                    values.difference[i],
                    values.difference[i] == 0,
                    values.similar[i],
                    values.difference[i] != 0 and not values.similar[i],
                    values.articles[i],
                    values.unit[i],
                    values.submodel[i],
                ]
                rows.append(row)
        return rows

    def export_excel(
        self,
        filepath: Path | str,
        sheets: list[Literal["extracted", "ignored", "definitions"]] | None = None,
        *,
        overwrite: bool = False,
    ) -> str | None:
        """Export evaluation data to an Excel file.

        Args:
            filepath (Path | str): The path where the Excel file will be saved.
            sheets (Literal["extracted", "ignored", "definitions"], optional):
                Specifies which sheets to include in the export.
                Options are "extracted", "ignored", and "definitions".
                Defaults to ["extracted"].
            overwrite (bool, optional): If True, existing file is overwritten.
                Defaults to False.

        Returns:
            str: The file path of the saved or existing Excel file. None in case
                of an error.

        """
        if sheets is None:
            sheets = ["extracted"]
        path = Path(filepath)
        if (not overwrite and path.exists()) or len(sheets) == 0:
            return str(path)
        workbook = openpyxl.Workbook()
        if workbook.active:
            workbook.remove(workbook.active)

        if "extracted" in sheets:
            self._add_table_sheet(
                workbook.create_sheet("extracted"),
                self.table_header,
                self._create_table(),
            )
        if "ignored" in sheets:
            self._add_table_sheet(
                workbook.create_sheet("ignored"),
                self.table_header,
                self._create_table(self.ignored_values),
            )
        if "definitions" in sheets:
            self._add_table_sheet(
                workbook.create_sheet("definitions"),
                self.table_header,
                self._create_definitions_table(),
            )

        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(path)
        except OSError:
            logger.exception("Couldn't save excel file to: %s", path)
            return None
        return str(path)

    def _add_table_sheet(self, sheet: Any, header: list, rows: list) -> None:
        sheet.append(header)
        for row in rows:
            sheet.append(row)
        sheet.auto_filter.ref = sheet.dimensions
        sheet.freeze_panes = sheet["C2"]

    def summary(self) -> str:
        """Return a summary string, with the evaluation and prompt counts."""
        defined_properties = sum(
            [len(article.definitions) if article.definitions else 0 for article in self.articles],
        )
        return f"""Evaluated {len(self.articles)} articles.
{defined_properties:3d} properties defined
{self.counts_sum.print() if self.counts_sum else ''}
{EvaluationPrompt.summarize(self.prompts)}"""

    def _print_values_filtered(
        self,
        property_filter: Callable | None = None,
        value_filter: Callable | None = None,
    ) -> str:
        print_string = ""
        for property_id, values in self.values.items():
            if property_filter and property_filter(property_id) is False:
                continue
            print_string += (
                f"\t{property_id} ({self.definitions[property_id].get_name(self.language)}):\n"
            )
            for i in range(len(values.extracted)):
                if value_filter and value_filter(values, i) is False:
                    continue
                print_string += f"\t\t{values.extracted[i]},\t {values.expected[i]},\t {values.difference[i]:.2f}\n"  # noqa: E501
        return print_string

    def log_values(self) -> None:
        """Log list of expected and extracted values."""
        logger.info(
            "Property values %s (extracted, expected, difference):\n%s",
            self.counts_sum.extracted,
            self._print_values_filtered(),
        )

    def log_correct(self) -> None:
        """Log list of correct expected and extracted values."""
        correct = self._print_values_filtered(
            property_filter=lambda property_id: self.counts[property_id].correct > 0,
            value_filter=lambda values, i: values.difference[i] == 0,
        )
        logger.info(
            "Correct property values %s (extracted, expected, difference):\n%s",
            self.counts_sum.correct,
            correct,
        )

    def log_different(self) -> None:
        """Log list of different expected and extracted values."""
        different = self._print_values_filtered(
            property_filter=lambda property_id: self.counts[property_id].different > 0,
            value_filter=lambda values, i: values.difference[i] != 0 and not values.similar[i],
        )
        logger.info(
            "Different property values %s (extracted, expected, difference):\n%s",
            self.counts_sum.different,
            different,
        )

    def log_similar(self) -> None:
        """Log list of similar expected and extracted values."""
        similar = self._print_values_filtered(
            property_filter=lambda property_id: self.counts[property_id].similar > 0,
            value_filter=lambda values, i: values.similar[i],
        )
        logger.info(
            "Similar property values %s (extracted, expected, difference):\n%s",
            self.counts_sum.similar,
            similar,
        )

    def plot_article_property_frequency(self, max_entries: int = 0) -> None:
        """Plot distribution of property counts in submodels."""
        _plot_discrete_hist(
            [
                property_.get_name(self.language)
                for article in self.articles
                for property_ in article.definitions
            ],
            limit=max_entries,
        )
        plt.title("Distribution of property counts in submodels")
        plt.ylabel("Property count")

    # def plot_article_property_frequency_eclass(self, max_entries=0):
    #     _plot_discrete_hist(
    #         [
    #             property.get_name(self.language)
    #             for article in self.articles
    #             if article.class_definition
    #             for property in article.class_definition.properties
    #         ],
    #         limit=max_entries,
    #     )
    #     plt.title("Distribution of property counts in eclass classes from datasheet")
    #     plt.ylabel("Property count")

    def plot_article_property_per_article(self, max_entries: int = 0) -> None:
        """Plot distribution of property count per article."""
        _plot_discrete_hist(
            [len(article.definitions) for article in self.articles],
            limit=max_entries,
            alpha=0.5,
        )
        # _plot_discrete_hist(
        #     [
        #         len(article.class_definition.properties) if article.class_definition else 0
        #         for article in self.articles
        #     ],
        #     limit=max_entries,
        #     alpha=0.5,
        # )
        # plt.legend(["submodel", "eclass (datasheet)"])
        plt.title("Distribution of property count per article")
        plt.xlabel("Property count per article")
        plt.ylabel("Article count")
        plt.grid(which="both", axis="both")

    # def plot_article_classes_datasheet(self, max_entries=0):
    #     bars, counts = _plot_discrete_hist(
    #         [
    #             article.class_definition.name if article.class_definition else "NONE"
    #             for article in self.articles
    #         ],
    #         limit=max_entries,
    #     )
    #     for bar, label in zip(bars, counts.keys()):
    #         if label == "NONE":
    #             bar.set_color("red")
    #             break
    #     plt.title(f"Distribution of ECLASS classes in datasheets")
    #     plt.ylabel("Article count")

    def plot_extraction_property_frequency(self, max_entries: int = 0) -> None:
        """Plot distribution of property counts with color-coded defined, correct etc."""
        self.plot_article_property_frequency(max_entries=max_entries)
        plt.title("Distribution of property counts")

        lables: list = [
            self.definitions[id_].get_name(self.language) if id_ in self.definitions else id_
            for id_ in self.counts
        ]
        plt.bar(lables, [count.ignored for count in self.counts.values()], color="purple")
        plt.bar(
            lables,
            [count.correct for count in self.counts.values()],
            bottom=[count.ignored for count in self.counts.values()],
            color="green",
        )
        plt.bar(
            lables,
            [count.similar for count in self.counts.values()],
            bottom=[count.ignored + count.correct for count in self.counts.values()],
            color="orange",
        )
        plt.bar(
            lables,
            [count.different for count in self.counts.values()],
            bottom=[
                count.ignored + count.correct + count.similar for count in self.counts.values()
            ],
            color="red",
        )
        plt.legend(["defined", "ignored", "correct", "similar", "different"])

    @staticmethod
    def plot_extraction_property_correct_similar_comparision(results: list["Evaluation"]) -> None:
        """Plot a comparision of correct incl. similar property counts for different results."""
        x_tick_set: set[str] = set()
        for result in results:
            x_tick_set.update(
                [result.definitions[id_].get_name("en", "Undefined") for id_ in result.counts],
            )
            x_tick_set.update(
                [
                    property_.get_name("en", "Undefined")
                    for article in result.articles
                    for property_ in article.definitions
                    if property_.get_name("en") not in Evaluation.ignore_properties
                ],
            )
        x_ticks = list(x_tick_set)
        x: list[float] = list(range(len(x_ticks)))

        bar_width = 1 / (len(results) + 1)
        for idx, result in enumerate(results):
            x = [x_i + bar_width for x_i in x]
            y1 = [0] * len(x)
            y2 = [0] * len(x)
            defined = Counter(
                [
                    property_.get_name("en")
                    for article in result.articles
                    for property_ in article.definitions
                    if property_.get_name("en") not in Evaluation.ignore_properties
                ],
            )
            for property_id, count in result.counts.items():
                property_name: str = result.definitions[property_id].get_name("en", "Undefined")
                if property_name in Evaluation.ignore_properties:
                    continue
                y1[x_ticks.index(property_name)] = defined[property_name]
                y2[x_ticks.index(property_name)] = count.correct + count.similar
            plt.bar(x, y1, width=bar_width, color="gray")
            plt.bar(x, y2, width=bar_width, label="Evaluation " + str(idx))

        plt.xlabel("Property")
        plt.xticks([x_i + 0.5 for x_i in list(range(len(x_ticks)))], x_ticks, rotation=90)
        plt.title("Comparision of correct incl. similar property counts")
        plt.ylabel("Property count (correct + similar)")
        plt.yticks()
        plt.grid(axis="both")
        plt.legend()

    @staticmethod
    def plot_extraction_property_correct_similar_comparision_boxplot(
        results: list["Evaluation"],
    ) -> None:
        """Plot a comparision of correct incl. similar property counts for different results.

        The results need to have same defined properties.
        """
        plt.boxplot(
            [
                [
                    (count.correct + count.similar) / count.compared
                    if count.compared > 0
                    else 0
                    for count in result.counts.values()
                ]
                for result in results
            ],
            label=["Evaluation " + str(i) for i in range(len(results))],
        )
        plt.title("Comparision of correct incl. similar property counts")
        plt.xlabel("Evaluation run")
        plt.ylabel("Relative Property count (correct + similar)")
        plt.grid(visible=True, axis="y")

    @staticmethod
    def plot_extraction_property_correct_similar_comparision_difference(
        result1: "Evaluation",
        result2: "Evaluation",
    ) -> None:
        """Plot a comparision of correct incl. similar property counts for two results.

        Creates a tree like bar chart, to compare which properties were extracted better.
        """
        counts: dict[str, int] = {}
        for property_id, count in result1.counts.items():
            name = (
                result1.definitions[property_id].get_name("en")
                if property_id in result1.definitions
                else property_id
            )
            counts[name] = -(count.correct + count.similar)
        for property_id, count in result2.counts.items():
            name = (
                result1.definitions[property_id].get_name("en")
                if property_id in result1.definitions
                else property_id
            )
            counts[name] = counts.get(name, 0) + (count.correct + count.similar)
        plt.barh(list(counts.keys()), list(counts.values()))

        plt.title(
            "Difference of correct incl. similar counts (- Evaluation 1 + Evaluation 2)",
        )
        plt.ylabel("Property count")
        plt.grid(axis="y")
